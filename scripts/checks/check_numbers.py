"""
check_numbers.py
----------------
Verify every inline number from numbers_manifest.xlsx appears correctly in the PRWP draft.

Handles Word tracked changes: includes <w:ins> text, excludes <w:del> text,
so edited numbers (e.g. 7,576 → 7,574) are read in their final accepted form.

Outputs:
  - PASS: stat found in expected paragraph
  - WARN: stat found in doc but not in expected paragraph
  - MISS: stat not found anywhere in the document
  - FLAG: manifest context string still contains a stale value

Run: python analyses/prwp_2025/scripts/check_numbers.py
"""
import os
import re
import openpyxl
from docx import Document
from lxml import etree

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(__file__)))
DOCX_PATH = os.path.join(REPO_ROOT, "manuscript", "PRWP working draft - Innovation Patterns in WBG Portfolio.docx")
MANIFEST_PATH = os.path.join(REPO_ROOT, "output", "numbers_manifest.xlsx")
SUMMARY_PATH = os.path.join(REPO_ROOT, "output", "numbers_summary.xlsx")

# XML namespaces
W_NS = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"


# ---------------------------------------------------------------------------
# Tracked-change-aware text extraction
# ---------------------------------------------------------------------------

def get_para_text_accepted(para_element) -> str:
    """
    Extract paragraph text as it would appear after accepting all tracked changes:
      - Include text inside <w:ins> elements
      - Exclude text inside <w:del> elements
      - Include plain <w:r> run text
    """
    parts = []
    for node in para_element.iter():
        tag = node.tag.split("}")[-1] if "}" in node.tag else node.tag

        # Skip deleted content
        if tag == "del":
            continue

        # Include text nodes that are direct children of runs (w:r) or insertions (w:ins > w:r)
        if tag == "t":
            # Check ancestors: if any ancestor is a w:del, skip
            parent = node.getparent()
            in_del = False
            while parent is not None:
                ptag = parent.tag.split("}")[-1] if "}" in parent.tag else parent.tag
                if ptag == "del":
                    in_del = True
                    break
                parent = parent.getparent()
            if not in_del and node.text:
                parts.append(node.text)
    return "".join(parts)


def build_doc_paragraphs(doc: Document) -> list[str]:
    """Return list of accepted-text strings, one per paragraph."""
    return [get_para_text_accepted(p._element) for p in doc.paragraphs]


# ---------------------------------------------------------------------------
# Number matching
# ---------------------------------------------------------------------------

def candidates(stat: str) -> list[str]:
    """
    Return a list of string forms to search for in the document text.
    Handles: % → "percent", comma-formatted integers, ~ > prefix, en-dash ranges,
    pp → "percentage point", × → "times", pts → "points".
    """
    s = str(stat).strip()
    variants = [s]

    # Strip leading modifier (~ > < ≈ ± + − -)
    core = re.sub(r"^[~><=≈±\+\-\u2212]+", "", s).strip()
    if core != s:
        variants.append(core)

    # Extract first bare number from stat (e.g. "0.13" from "+0.13 pts (3.2%)")
    first_num = re.search(r"\d[\d.,]*", core)
    if first_num:
        variants.append(first_num.group())

    # Numeric with % (simple or with modifier): try "X percent"
    m = re.match(r"^[~><=≈±]?([0-9.,]+)%$", s)
    if m:
        n = m.group(1)
        variants += [f"{n} percent", f"{n} per cent"]

    # Range with % like "40–46%" — try each bound separately
    m = re.match(r"^([0-9.]+)[–\-]([0-9.]+)%$", s)
    if m:
        lo, hi = m.group(1), m.group(2)
        variants += [
            f"{lo} percent", f"{hi} percent",
            f"{lo}–{hi} percent", f"{lo}-{hi} percent",
        ]

    # pp → percentage points
    if "pp" in s:
        core_pp = re.sub(r"pp", "", s).strip().lstrip("+ \u2212-")
        variants += [core_pp, f"{core_pp} percentage point", f"{core_pp} pp"]

    # × → "times"
    if "×" in s:
        n_x = re.sub(r"[~×]", "", s).strip()
        variants += [f"{n_x} times", f"{n_x}x", f"{n_x}-fold"]

    # pts → points (handle unicode minus and parenthetical like "+0.13 pts (3.2%)")
    m = re.search(r"([0-9.,]+)\s*pts", s)
    if m:
        n_pts = m.group(1)
        variants += [f"{n_pts} point", f"{n_pts} points"]

    # Comma-formatted integers only (4+ digit sequences, no decimal)
    int_only = re.match(r"^[~><=≈±\+\-\u2212]?(\d{4,})$", s)
    if int_only:
        n = int(int_only.group(1))
        variants += [f"{n:,}", str(n)]

    # Deduplicate, preserving order
    seen, result = set(), []
    for v in variants:
        vl = v.lower().strip()
        if vl and vl not in seen:
            seen.add(vl)
            result.append(vl)
    return result


def stat_in_text(stat: str, text: str) -> bool:
    """Check if any candidate form of stat appears in text (case-insensitive)."""
    text_lower = text.lower()
    return any(c in text_lower for c in candidates(stat))


def load_manifest() -> list[dict]:
    wb = openpyxl.load_workbook(MANIFEST_PATH)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    cols = rows[0]
    return [dict(zip(cols, row)) for row in rows[1:]]


def load_summary() -> dict[str, str]:
    wb = openpyxl.load_workbook(SUMMARY_PATH)
    ws = wb.active
    return {str(row[0]): str(row[1]) for row in ws.iter_rows(values_only=True) if row[0]}


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    doc = Document(DOCX_PATH)
    paragraphs = build_doc_paragraphs(doc)
    manifest = load_manifest()
    summary = load_summary()

    passes, warns, misses, flags = [], [], [], []

    print("=" * 70)
    print("NUMBER CHECK: manifest vs PRWP draft")
    print("=" * 70)

    for entry in manifest:
        stat = str(entry["stat"])
        context = entry.get("context") or ""
        para_idx = entry.get("para_index")
        table_ref = entry.get("table_ref")

        # Check if stat is in the expected paragraph
        if para_idx is not None and para_idx < len(paragraphs):
            expected_para = paragraphs[para_idx]
            if stat_in_text(stat, expected_para):
                passes.append((stat, context[:60]))
                continue

        # Not found in expected para — search whole doc
        found_at = [i for i, p in enumerate(paragraphs) if stat_in_text(stat, p)]
        if found_at:
            warns.append((stat, para_idx, found_at, context[:60]))
        else:
            misses.append((stat, para_idx, context[:60]))

    # Check manifest context strings for stale values
    stale_patterns = ["7,576"]  # add more if needed
    for entry in manifest:
        context = entry.get("context") or ""
        for stale in stale_patterns:
            if stale in context:
                flags.append((entry["stat"], stale, context[:80]))

    # --- Report ---
    print(f"\nPASS  {len(passes):>3} — stat found in expected paragraph")
    print(f"WARN  {len(warns):>3} — stat found in doc but not at expected paragraph index")
    print(f"MISS  {len(misses):>3} — stat not found anywhere in document")
    print(f"FLAG  {len(flags):>3} — manifest context string contains stale value")

    if warns:
        print("\n--- WARN: found at wrong paragraph ---")
        for stat, expected, found_at, ctx in warns:
            print(f"  [{stat}] expected para {expected}, found at {found_at}")
            print(f"    context: {ctx}")

    if misses:
        print("\n--- MISS: stat not found in document ---")
        for stat, expected, ctx in misses:
            print(f"  [{stat}] (expected para {expected})")
            print(f"    context: {ctx}")

    if flags:
        print("\n--- FLAG: stale context strings in manifest ---")
        for stat, stale, ctx in flags:
            print(f"  stat={stat} | stale='{stale}' in: {ctx}")

    # Cross-check summary vs manifest for Total Projects
    print("\n--- Cross-check: numbers_summary.xlsx vs manifest ---")
    summary_total = summary.get("Total Projects", "?")
    manifest_total = next(
        (str(e["stat"]) for e in manifest if "7,57" in str(e["stat"]) and "project-level dataset" in (e.get("context") or "")),
        "not found"
    )
    match = "✓" if str(summary_total) in manifest_total.replace(",", "") else "✗"
    print(f"  numbers_summary Total Projects: {summary_total}")
    print(f"  manifest stat for that entry:   {manifest_total}")
    print(f"  Match: {match}")

    if len(passes) == len(manifest):
        print("\nAll numbers verified ✓")


if __name__ == "__main__":
    main()
